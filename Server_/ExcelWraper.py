import openpyxl
from peewee import *
from Classes import Product

def GetHeaderByNumCol(headres_row:int,filename:str,sheet_name:str):
    # * ПОЛУЧЕНИЕ  ЗАГОЛОВКА ПО НОМЕРУ КОЛОНКИ
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    
    headers = {}
    for idx,cell in enumerate(sheet[int(headres_row)]):
        headers[idx]=cell.value
        
    return headers


def GetNumColByHeader(headres_row:int,filename:str,sheet_name:str):
    # * ПОЛУЧЕНИЕ НОМЕРА КОЛОНКИ ПО ЗАГОЛОВКУ   
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    
    headers = {}
    for idx,cell in enumerate(sheet[int(headres_row)]):
        headers[cell.value]=idx
        
    return headers

def InsertByHeaderName(header_name,filename,headers_dict,sheet_name):
    # ! NOT WORK
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    
def GetDataByHeaders(target_idxs:list,header_idxs:dict,filename:str,sheet_name:str,begin_row:int):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    
    res = []
    row = sheet.iter_rows(min_row=begin_row, values_only=True)
    for idx,cell in enumerate(row):
        r = {}
        for ind,val in enumerate(cell):
            if ind in target_idxs:
                r[header_idxs[ind]]=val
        res.append(r)
            
    return res
            
def InsertDataFromExcel_ToDatabase(filename:str,headers_row:int,sheet_name:str):
    NumColByHeaders = GetNumColByHeader(filename=filename,headres_row=headers_row,sheet_name=sheet_name)
    HeadersByNumCol = GetHeaderByNumCol(filename=filename,headres_row=headers_row,sheet_name=sheet_name)

    # GET DATA FROM EXCEL
    target_idxs = [
        NumColByHeaders['Артикул*'],
        NumColByHeaders['Название товара'],
        NumColByHeaders['Цена, руб.*'],
        NumColByHeaders['Ссылки на дополнительные фото'],
    ]
    r = GetDataByHeaders(filename='test.xlsx',header_idxs=HeadersByNumCol,sheet_name='Шаблон',begin_row=4,target_idxs=target_idxs)

    # print(r)



    # define the database
    db = SqliteDatabase('DB.db')

   

    db.create_tables([Product])
    # connect to the database
    # db.connect()

    # todo INSERT DATA
    for el in r:
        data = {'article':el['Артикул*'],'name':el['Название товара'],'price':el['Цена, руб.*'],'imgs':el['Ссылки на дополнительные фото']}
        
        # Product.insert(data).on_conflict(conflict_target=[Product.article],preserve=[Product.name,Product.imgs,Product.price]).execute()
        Product.insert(data).on_conflict_replace(replace=True).execute()

    db.close()
    print('Данные вставлены')

InsertDataFromExcel_ToDatabase(filename='test.xlsx',headers_row=2,sheet_name='Шаблон')
# for i in Product.select(Product.article):
#     print(i.article)

